"""Unit and integration tests for AdMetric Pro main module.

Tests the CLI interface and end-to-end pipeline from CSV input
to formatted Excel report output.
"""

import os
import tempfile
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.main import parse_arguments, main


class TestParseArguments:
    """Tests for CLI argument parsing.

    Validates that the argparse configuration correctly handles
    all supported command-line arguments for AdMetric Pro.
    """

    def test_parse_required_input_file(self) -> None:
        """Test that input file argument is correctly parsed.

        Returns:
            None. Asserts input_file matches expected value.
        """
        args = parse_arguments(["data/meta_ads.csv"])
        assert args.input_file == "data/meta_ads.csv"

    def test_parse_default_output_directory(self) -> None:
        """Test default output directory is 'output'.

        Returns:
            None. Asserts output defaults to 'output'.
        """
        args = parse_arguments(["data/meta_ads.csv"])
        assert args.output == "output"

    def test_parse_custom_output_directory(self) -> None:
        """Test custom output directory with --output flag.

        Returns:
            None. Asserts custom output path is captured.
        """
        args = parse_arguments(["data/meta_ads.csv", "--output", "reports/"])
        assert args.output == "reports/"

    def test_parse_output_short_flag(self) -> None:
        """Test custom output directory with -o short flag.

        Returns:
            None. Asserts short flag works identically to long form.
        """
        args = parse_arguments(["data/meta_ads.csv", "-o", "custom_output"])
        assert args.output == "custom_output"

    def test_parse_default_cpc_threshold(self) -> None:
        """Test default CPC threshold is 20.00.

        Returns:
            None. Asserts CPC threshold defaults to R20.00.
        """
        args = parse_arguments(["data/meta_ads.csv"])
        assert args.cpc_threshold == 20.00

    def test_parse_custom_cpc_threshold(self) -> None:
        """Test custom CPC threshold with --cpc-threshold flag.

        Returns:
            None. Asserts custom threshold is captured.
        """
        args = parse_arguments(["data/meta_ads.csv", "--cpc-threshold", "25.50"])
        assert args.cpc_threshold == 25.50

    def test_parse_cpc_threshold_short_flag(self) -> None:
        """Test custom CPC threshold with -t short flag.

        Returns:
            None. Asserts short flag works identically to long form.
        """
        args = parse_arguments(["data/meta_ads.csv", "-t", "15.00"])
        assert args.cpc_threshold == 15.00

    def test_parse_verbose_flag(self) -> None:
        """Test verbose flag is correctly parsed.

        Returns:
            None. Asserts verbose is True when flag provided.
        """
        args = parse_arguments(["data/meta_ads.csv", "--verbose"])
        assert args.verbose is True

    def test_parse_verbose_short_flag(self) -> None:
        """Test verbose short flag -v.

        Returns:
            None. Asserts short flag works identically to long form.
        """
        args = parse_arguments(["data/meta_ads.csv", "-v"])
        assert args.verbose is True

    def test_parse_default_verbose_is_false(self) -> None:
        """Test verbose defaults to False.

        Returns:
            None. Asserts verbose is False when flag omitted.
        """
        args = parse_arguments(["data/meta_ads.csv"])
        assert args.verbose is False

    def test_parse_all_arguments_combined(self) -> None:
        """Test parsing all arguments together.

        Returns:
            None. Asserts all arguments are correctly captured.
        """
        args = parse_arguments([
            "input.csv",
            "--output", "reports",
            "--cpc-threshold", "30.00",
            "--verbose"
        ])
        assert args.input_file == "input.csv"
        assert args.output == "reports"
        assert args.cpc_threshold == 30.00
        assert args.verbose is True


class TestMainIntegration:
    """Integration tests for the complete pipeline.

    Validates end-to-end report generation from CSV input to
    formatted Excel output, including sheet creation and metric
    calculations.
    """

    @pytest.fixture
    def sample_csv_data(self) -> dict[str, list]:
        """Mock Meta Ads CSV data for testing.

        Returns:
            Dictionary with campaign data matching Meta Ads export format.
            Includes campaigns with known CTR/CPC values for verification.
        """
        return {
            "Campaign Name": ["Summer Sale", "Winter Promo", "Brand Awareness"],
            "Amount Spent (ZAR)": [1000.00, 500.00, 2500.00],
            "Link Clicks": [50, 25, 100],
            "Impressions": [2500, 1250, 10000],
        }

    @pytest.fixture
    def temp_csv_file(self, sample_csv_data: dict[str, list]) -> str:
        """Create a temporary CSV file with sample data.

        Args:
            sample_csv_data: Dictionary of campaign data to write.

        Yields:
            Path to the temporary CSV file.
        """
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            df = pd.DataFrame(sample_csv_data)
            df.to_csv(f.name, index=False)
            yield f.name
        os.unlink(f.name)

    @pytest.fixture
    def temp_output_dir(self) -> str:
        """Create a temporary output directory.

        Yields:
            Path to the temporary directory.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    def test_main_success_generates_report(
        self, temp_csv_file: str, temp_output_dir: str
    ) -> None:
        """Test successful end-to-end report generation.

        Args:
            temp_csv_file: Path to test CSV file.
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts exit code is 0 and Excel file is created.
        """
        with patch("sys.argv", ["main", temp_csv_file, "--output", temp_output_dir]):
            exit_code = main()

        assert exit_code == 0

        # Verify Excel file was created
        output_files = list(Path(temp_output_dir).glob("*.xlsx"))
        assert len(output_files) == 1
        assert "AdMetric_Pro_Report_" in output_files[0].name

    def test_main_report_contains_campaign_details(
        self, temp_csv_file: str, temp_output_dir: str
    ) -> None:
        """Test that generated report contains Campaign Details sheet.

        Args:
            temp_csv_file: Path to test CSV file.
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts Campaign Details sheet exists.
        """
        with patch("sys.argv", ["main", temp_csv_file, "--output", temp_output_dir]):
            main()

        output_file = list(Path(temp_output_dir).glob("*.xlsx"))[0]
        wb = load_workbook(output_file)

        assert "Campaign Details" in wb.sheetnames

    def test_main_report_contains_executive_summary(
        self, temp_csv_file: str, temp_output_dir: str
    ) -> None:
        """Test that generated report contains Executive Summary sheet.

        Args:
            temp_csv_file: Path to test CSV file.
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts Executive Summary sheet exists.
        """
        with patch("sys.argv", ["main", temp_csv_file, "--output", temp_output_dir]):
            main()

        output_file = list(Path(temp_output_dir).glob("*.xlsx"))[0]
        wb = load_workbook(output_file)

        assert "Executive Summary" in wb.sheetnames

    def test_main_ctr_calculation_correct(
        self, temp_csv_file: str, temp_output_dir: str
    ) -> None:
        """Test CTR is calculated correctly: (clicks/impressions)*100.

        Verifies Summer Sale campaign: 50 clicks / 2500 impressions = 2.0%.

        Args:
            temp_csv_file: Path to test CSV file.
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts CTR value matches expected calculation.
        """
        with patch("sys.argv", ["main", temp_csv_file, "--output", temp_output_dir]):
            main()

        output_file = list(Path(temp_output_dir).glob("*.xlsx"))[0]
        wb = load_workbook(output_file)
        ws = wb["Campaign Details"]

        # Row 2 is first data row (Summer Sale: 50 clicks / 2500 impressions = 2.0%)
        ctr_value = ws.cell(row=2, column=5).value  # CTR column
        assert ctr_value == 2.0

    def test_main_cpc_calculation_correct(
        self, temp_csv_file: str, temp_output_dir: str
    ) -> None:
        """Test CPC is calculated correctly: spend/clicks.

        Verifies Summer Sale campaign: R1000 / 50 clicks = R20.00.

        Args:
            temp_csv_file: Path to test CSV file.
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts CPC value matches expected calculation.
        """
        with patch("sys.argv", ["main", temp_csv_file, "--output", temp_output_dir]):
            main()

        output_file = list(Path(temp_output_dir).glob("*.xlsx"))[0]
        wb = load_workbook(output_file)
        ws = wb["Campaign Details"]

        # Row 2 is first data row (Summer Sale: R1000 / 50 clicks = R20.00)
        cpc_value = ws.cell(row=2, column=6).value  # CPC column
        assert cpc_value == 20.0

    def test_main_file_not_found_returns_error(self, temp_output_dir: str) -> None:
        """Test that missing input file returns exit code 1.

        Args:
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts exit code is 1 for missing file.
        """
        with patch("sys.argv", ["main", "nonexistent.csv", "--output", temp_output_dir]):
            exit_code = main()

        assert exit_code == 1

    def test_main_invalid_csv_returns_error(self, temp_output_dir: str) -> None:
        """Test that invalid CSV returns exit code 1.

        Args:
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts exit code is 1 for invalid CSV structure.
        """
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            f.write("Invalid,Headers,Only\n1,2,3\n")
            invalid_csv = f.name

        try:
            with patch("sys.argv", ["main", invalid_csv, "--output", temp_output_dir]):
                exit_code = main()
            assert exit_code == 1
        finally:
            os.unlink(invalid_csv)

    def test_main_empty_csv_returns_error(self, temp_output_dir: str) -> None:
        """Test that empty CSV (headers only, no data) returns exit code 1.

        Args:
            temp_output_dir: Path to output directory.

        Returns:
            None. Asserts exit code is 1 for empty data.
        """
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            f.write("Campaign Name,Amount Spent (ZAR),Link Clicks,Impressions\n")
            empty_csv = f.name

        try:
            with patch("sys.argv", ["main", empty_csv, "--output", temp_output_dir]):
                exit_code = main()
            assert exit_code == 1
        finally:
            os.unlink(empty_csv)


class TestMainMetricsVerification:
    """Tests verifying metric calculations match expected values.

    Focuses on edge cases like zero-division protection and
    high CPC threshold detection for red flag highlighting.
    """

    @pytest.fixture
    def edge_case_csv_data(self) -> dict[str, list]:
        """CSV data with edge cases: zero clicks, high CPC.

        Returns:
            Dictionary with campaigns designed to test edge cases:
            - Zero Clicks: Tests zero-division protection for CPC/CTR.
            - High CPC: Tests R50 CPC calculation (above R20 threshold).
            - Normal: Baseline campaign for comparison.
        """
        return {
            "Campaign Name": ["Zero Clicks", "High CPC", "Normal"],
            "Amount Spent (ZAR)": [100.00, 500.00, 200.00],
            "Link Clicks": [0, 10, 20],
            "Impressions": [1000, 500, 4000],
        }

    @pytest.fixture
    def edge_case_csv_file(self, edge_case_csv_data: dict[str, list]) -> str:
        """Create CSV with edge case data.

        Args:
            edge_case_csv_data: Dictionary of edge case campaign data.

        Yields:
            Path to the temporary CSV file.
        """
        with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False) as f:
            df = pd.DataFrame(edge_case_csv_data)
            df.to_csv(f.name, index=False)
            yield f.name
        os.unlink(f.name)

    def test_zero_clicks_cpc_is_zero(self, edge_case_csv_file: str) -> None:
        """Test CPC is 0 when clicks are 0 (zero-division protection).

        Verifies Zero Clicks campaign: R100 / 0 clicks = R0.00 (protected).

        Args:
            edge_case_csv_file: Path to edge case CSV file.

        Returns:
            None. Asserts CPC is 0.0 instead of raising ZeroDivisionError.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("sys.argv", ["main", edge_case_csv_file, "--output", tmpdir]):
                main()

            output_file = list(Path(tmpdir).glob("*.xlsx"))[0]
            wb = load_workbook(output_file)
            ws = wb["Campaign Details"]

            # Row 2: Zero Clicks campaign
            cpc_value = ws.cell(row=2, column=6).value
            assert cpc_value == 0.0

    def test_high_cpc_value_calculated(self, edge_case_csv_file: str) -> None:
        """Test high CPC is calculated correctly: R500/10 = R50.

        Verifies High CPC campaign exceeds R20 threshold for red flagging.

        Args:
            edge_case_csv_file: Path to edge case CSV file.

        Returns:
            None. Asserts CPC value is R50.00.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("sys.argv", ["main", edge_case_csv_file, "--output", tmpdir]):
                main()

            output_file = list(Path(tmpdir).glob("*.xlsx"))[0]
            wb = load_workbook(output_file)
            ws = wb["Campaign Details"]

            # Row 3: High CPC campaign (R500 / 10 clicks = R50)
            cpc_value = ws.cell(row=3, column=6).value
            assert cpc_value == 50.0

    def test_ctr_with_zero_clicks(self, edge_case_csv_file: str) -> None:
        """Test CTR is 0% when clicks are 0.

        Verifies Zero Clicks campaign: 0 / 1000 impressions = 0.0%.

        Args:
            edge_case_csv_file: Path to edge case CSV file.

        Returns:
            None. Asserts CTR is 0.0 instead of raising ZeroDivisionError.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("sys.argv", ["main", edge_case_csv_file, "--output", tmpdir]):
                main()

            output_file = list(Path(tmpdir).glob("*.xlsx"))[0]
            wb = load_workbook(output_file)
            ws = wb["Campaign Details"]

            # Row 2: Zero Clicks campaign (0/1000 = 0%)
            ctr_value = ws.cell(row=2, column=5).value
            assert ctr_value == 0.0
