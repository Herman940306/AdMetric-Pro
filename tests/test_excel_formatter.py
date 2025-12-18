"""Unit tests for the Excel Formatter module.

Tests cover header formatting, currency formatting, conditional formatting,
Executive Summary creation, and report generation.
"""

import pytest
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.excel_formatter import (
    generate_timestamped_filename,
    apply_header_formatting,
    apply_currency_formatting,
    apply_conditional_formatting,
    create_executive_summary,
    generate_report,
    HEADER_FILL,
    HIGH_CPC_FILL,
    CPC_THRESHOLD,
)


class TestGenerateTimestampedFilename:
    """Test suite for generate_timestamped_filename function."""

    def test_filename_format(self) -> None:
        """Test that filename follows expected format."""
        filename = generate_timestamped_filename()
        assert filename.startswith("AdMetric_Pro_Report_")
        assert filename.endswith(".xlsx")

    def test_filename_contains_date(self) -> None:
        """Test that filename contains current date."""
        filename = generate_timestamped_filename()
        today = datetime.now().strftime("%Y-%m-%d")
        assert today in filename

    def test_filename_is_unique_per_minute(self) -> None:
        """Test that filename includes time component."""
        filename = generate_timestamped_filename()
        # Format: AdMetric_Pro_Report_YYYY-MM-DD_HHMM.xlsx
        parts = filename.replace(".xlsx", "").split("_")
        assert len(parts) == 5  # AdMetric, Pro, Report, date, time


class TestApplyHeaderFormatting:
    """Test suite for apply_header_formatting function."""

    def test_header_has_dark_blue_fill(self, tmp_path: Path) -> None:
        """Test that headers have dark blue background."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Campaign Name"
        ws["B1"] = "Amount Spent"

        apply_header_formatting(ws)

        assert ws["A1"].fill.start_color.rgb == "001F4E79"
        assert ws["B1"].fill.start_color.rgb == "001F4E79"

    def test_header_has_white_bold_font(self) -> None:
        """Test that headers have white bold text."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test Header"

        apply_header_formatting(ws)

        assert ws["A1"].font.bold is True
        assert ws["A1"].font.color.rgb == "00FFFFFF"


class TestApplyConditionalFormatting:
    """Test suite for apply_conditional_formatting function."""

    def test_high_cpc_rows_highlighted(self, tmp_path: Path) -> None:
        """Test that rows with CPC > R20.00 are highlighted red."""
        from openpyxl import Workbook

        df = pd.DataFrame({
            "campaign_name": ["Low CPC", "High CPC"],
            "amount_spent": [100.0, 500.0],
            "link_clicks": [10, 10],
            "impressions": [1000, 1000],
            "ctr": [1.0, 1.0],
            "cpc": [10.0, 50.0]  # Second row exceeds threshold
        })

        wb = Workbook()
        ws = wb.active

        # Write data
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        apply_conditional_formatting(ws, "cpc", df)

        # Row 2 (Low CPC) should NOT be highlighted
        assert ws.cell(row=2, column=1).fill.start_color.rgb != "00FFC7CE"
        # Row 3 (High CPC) SHOULD be highlighted
        assert ws.cell(row=3, column=1).fill.start_color.rgb == "00FFC7CE"

    def test_threshold_boundary(self) -> None:
        """Test that exactly R20.00 is NOT highlighted (only > 20)."""
        from openpyxl import Workbook

        df = pd.DataFrame({
            "campaign_name": ["Exactly 20"],
            "cpc": [20.0]
        })

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "campaign_name"
        ws["B1"] = "cpc"
        ws["A2"] = "Exactly 20"
        ws["B2"] = 20.0

        apply_conditional_formatting(ws, "cpc", df)

        # Exactly R20.00 should NOT be highlighted
        assert ws.cell(row=2, column=1).fill.start_color.rgb != "00FFC7CE"


class TestCreateExecutiveSummary:
    """Test suite for create_executive_summary function."""

    def test_summary_sheet_created(self) -> None:
        """Test that Executive Summary sheet is created."""
        from openpyxl import Workbook

        df = pd.DataFrame({
            "campaign_name": ["Test"],
            "amount_spent": [1000.0],
            "link_clicks": [50],
            "impressions": [2500],
            "ctr": [2.0],
            "cpc": [20.0]
        })

        wb = Workbook()
        summary_ws = create_executive_summary(wb, df)

        assert summary_ws.title == "Executive Summary"

    def test_summary_contains_total_spend(self) -> None:
        """Test that summary displays correct Total Spend."""
        from openpyxl import Workbook

        df = pd.DataFrame({
            "campaign_name": ["A", "B"],
            "amount_spent": [1000.0, 2000.0],
            "link_clicks": [50, 100],
            "impressions": [2500, 5000],
            "ctr": [2.0, 2.0],
            "cpc": [20.0, 20.0]
        })

        wb = Workbook()
        summary_ws = create_executive_summary(wb, df)

        # Find Total Spend value
        found_spend = False
        for row in summary_ws.iter_rows(min_row=1, max_row=10, max_col=2):
            for cell in row:
                if cell.value and "3,000.00" in str(cell.value):
                    found_spend = True
                    break

        assert found_spend, "Total Spend R3,000.00 not found in summary"

    def test_summary_handles_zero_clicks(self) -> None:
        """Test that summary handles zero clicks without division error."""
        from openpyxl import Workbook

        df = pd.DataFrame({
            "campaign_name": ["Zero Clicks"],
            "amount_spent": [500.0],
            "link_clicks": [0],
            "impressions": [1000],
            "ctr": [0.0],
            "cpc": [0.0]
        })

        wb = Workbook()
        # Should not raise any errors
        summary_ws = create_executive_summary(wb, df)
        assert summary_ws is not None


class TestGenerateReport:
    """Test suite for generate_report function."""

    def test_report_creates_file(self, tmp_path: Path) -> None:
        """Test that report generates an Excel file."""
        df = pd.DataFrame({
            "campaign_name": ["Test Campaign"],
            "amount_spent": [1000.0],
            "link_clicks": [50],
            "impressions": [2500],
            "ctr": [2.0],
            "cpc": [20.0]
        })

        report_path = generate_report(df, str(tmp_path))

        assert Path(report_path).exists()
        assert report_path.endswith(".xlsx")

    def test_report_has_two_sheets(self, tmp_path: Path) -> None:
        """Test that report contains Campaign Details and Executive Summary sheets."""
        df = pd.DataFrame({
            "campaign_name": ["Test"],
            "amount_spent": [1000.0],
            "link_clicks": [50],
            "impressions": [2500],
            "ctr": [2.0],
            "cpc": [20.0]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)

        assert "Campaign Details" in wb.sheetnames
        assert "Executive Summary" in wb.sheetnames

    def test_report_preserves_data(self, tmp_path: Path) -> None:
        """Test that report preserves campaign data correctly."""
        df = pd.DataFrame({
            "campaign_name": ["Summer Sale"],
            "amount_spent": [1500.0],
            "link_clicks": [75],
            "impressions": [5000],
            "ctr": [1.5],
            "cpc": [20.0]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Check header row
        assert ws["A1"].value == "Campaign Name"
        # Check data row
        assert ws["A2"].value == "Summer Sale"

    def test_report_applies_conditional_formatting(self, tmp_path: Path) -> None:
        """Test that high CPC campaigns are highlighted."""
        df = pd.DataFrame({
            "campaign_name": ["Low CPC", "High CPC"],
            "amount_spent": [100.0, 500.0],
            "link_clicks": [10, 10],
            "impressions": [1000, 1000],
            "ctr": [1.0, 1.0],
            "cpc": [10.0, 50.0]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Row 3 (High CPC) should be highlighted
        assert ws.cell(row=3, column=1).fill.start_color.rgb == "00FFC7CE"

    def test_report_creates_output_directory(self, tmp_path: Path) -> None:
        """Test that report creates output directory if it doesn't exist."""
        df = pd.DataFrame({
            "campaign_name": ["Test"],
            "amount_spent": [1000.0],
            "link_clicks": [50],
            "impressions": [2500],
            "ctr": [2.0],
            "cpc": [20.0]
        })

        new_dir = tmp_path / "new_output_dir"
        report_path = generate_report(df, str(new_dir))

        assert new_dir.exists()
        assert Path(report_path).exists()
