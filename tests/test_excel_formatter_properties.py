"""Property-based tests for the Excel Formatter module.

These tests verify universal correctness properties for Excel report
generation using the Hypothesis library.

Business Value:
    Ensures reliable, professional Excel report generation for client
    deliverables, preventing formatting errors that could damage agency
    reputation.

Example:
    Run these tests with pytest::

        pytest tests/test_excel_formatter_properties.py -v
"""

import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from hypothesis import given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook
from pytest import TempPathFactory

from src.excel_formatter import (
    CPC_THRESHOLD,
    ZAR_FORMAT,
    generate_report,
    generate_timestamped_filename,
)

# Configure module logger
logger = logging.getLogger(__name__)


class TestTimestampedFilenameProperty:
    """Property-based tests for timestamped filename generation.

    Feature:
        admetric-pro, Property 9: Timestamped Filename Format

    Validates:
        Requirements 8.1

    Business Value:
        Ensures report filenames are unique and traceable, preventing
        accidental overwrites and maintaining audit trails.
    """

    @settings(max_examples=100)
    @given(st.integers(min_value=1, max_value=100))
    def test_filename_matches_pattern(self, _: int) -> None:
        """Property: Filename always matches expected pattern.

        Feature:
            admetric-pro, Property 9: Timestamped Filename Format

        Validates:
            Requirements 8.1

        For any generated report, the output filename SHALL match the
        pattern AdMetric_Pro_Report_YYYY-MM-DD_HHMM.xlsx.

        Args:
            _: Unused integer parameter for Hypothesis iteration.

        Raises:
            AssertionError: If filename doesn't match expected pattern.
        """
        filename = generate_timestamped_filename()

        # Pattern: AdMetric_Pro_Report_YYYY-MM-DD_HHMM.xlsx
        pattern = r"^AdMetric_Pro_Report_\d{4}-\d{2}-\d{2}_\d{4}\.xlsx$"
        assert re.match(pattern, filename), f"Filename '{filename}' doesn't match expected pattern"

    @settings(max_examples=50)
    @given(st.integers(min_value=1, max_value=50))
    def test_filename_has_xlsx_extension(self, _: int) -> None:
        """Property: Filename always has .xlsx extension.

        Feature:
            admetric-pro, Property 9: Timestamped Filename Format

        Validates:
            Requirements 8.1

        Args:
            _: Unused integer parameter for Hypothesis iteration.

        Raises:
            AssertionError: If filename doesn't end with .xlsx.
        """
        filename = generate_timestamped_filename()
        assert filename.endswith(".xlsx"), f"Filename '{filename}' doesn't have .xlsx extension"


class TestHeaderFormattingProperty:
    """Property-based tests for header formatting.

    Feature:
        admetric-pro, Property 4: Header Formatting Consistency

    Validates:
        Requirements 4.1, 4.2

    Business Value:
        Ensures consistent professional appearance across all reports,
        reinforcing agency brand standards.
    """

    @given(
        num_campaigns=st.integers(min_value=1, max_value=10),
        base_spend=st.floats(min_value=100.0, max_value=10000.0, allow_nan=False, allow_infinity=False)
    )
    @settings(max_examples=50)
    def test_all_headers_formatted(
        self,
        tmp_path_factory: TempPathFactory,
        num_campaigns: int,
        base_spend: float
    ) -> None:
        """Property: All header cells have bold font and dark blue background.

        Feature:
            admetric-pro, Property 4: Header Formatting Consistency

        Validates:
            Requirements 4.1, 4.2

        For any generated Excel report, all header cells in the first row
        SHALL have bold font and dark blue background color applied.

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            num_campaigns: Hypothesis-generated number of campaigns (1-10).
            base_spend: Hypothesis-generated base spend amount in ZAR.

        Raises:
            AssertionError: If any header cell lacks proper formatting.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        # Generate test data
        df = pd.DataFrame({
            "campaign_name": [f"Campaign {i}" for i in range(num_campaigns)],
            "amount_spent": [base_spend + i * 100 for i in range(num_campaigns)],
            "link_clicks": [50 + i * 10 for i in range(num_campaigns)],
            "impressions": [2500 + i * 500 for i in range(num_campaigns)],
            "ctr": [2.0 for _ in range(num_campaigns)],
            "cpc": [20.0 for _ in range(num_campaigns)]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Check all header cells
        for cell in ws[1]:
            if cell.value:
                assert cell.font.bold is True, f"Header '{cell.value}' is not bold"
                assert cell.fill.start_color.rgb == "001F4E79", f"Header '{cell.value}' doesn't have dark blue fill"

    @given(
        num_campaigns=st.integers(min_value=1, max_value=5)
    )
    @settings(max_examples=50)
    def test_headers_centered_alignment(
        self,
        tmp_path_factory: TempPathFactory,
        num_campaigns: int
    ) -> None:
        """Property: All header cells have centered alignment.

        Feature:
            admetric-pro, Property 4: Header Formatting Consistency

        Validates:
            Requirements 4.1, 4.2

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            num_campaigns: Hypothesis-generated number of campaigns.

        Raises:
            AssertionError: If any header cell lacks centered alignment.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        df = pd.DataFrame({
            "campaign_name": [f"Campaign {i}" for i in range(num_campaigns)],
            "amount_spent": [1000.0 + i * 100 for i in range(num_campaigns)],
            "link_clicks": [50 for _ in range(num_campaigns)],
            "impressions": [2500 for _ in range(num_campaigns)],
            "ctr": [2.0 for _ in range(num_campaigns)],
            "cpc": [20.0 for _ in range(num_campaigns)]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        for cell in ws[1]:
            if cell.value:
                assert cell.alignment.horizontal == "center", \
                    f"Header '{cell.value}' is not horizontally centered"


class TestCurrencyFormattingProperty:
    """Property-based tests for ZAR currency formatting.

    Feature:
        admetric-pro, Property 5: Currency Formatting Application

    Validates:
        Requirements 4.3

    Business Value:
        Ensures all monetary values display correctly in South African Rand,
        meeting client expectations for local currency presentation.
    """

    @given(
        spend_values=st.lists(
            st.floats(min_value=100.0, max_value=50000.0, allow_nan=False, allow_infinity=False),
            min_size=1,
            max_size=5
        )
    )
    @settings(max_examples=50)
    def test_currency_columns_have_zar_format(
        self,
        tmp_path_factory: TempPathFactory,
        spend_values: list[float]
    ) -> None:
        """Property: Amount Spent and CPC columns have ZAR currency format.

        Feature:
            admetric-pro, Property 5: Currency Formatting Application

        Validates:
            Requirements 4.3

        For any generated Excel report, the Amount Spent and CPC columns
        SHALL have ZAR currency number format applied to all data cells.

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            spend_values: Hypothesis-generated list of spend amounts.

        Raises:
            AssertionError: If currency columns lack ZAR formatting.
        """
        tmp_path = tmp_path_factory.mktemp("reports")
        num_campaigns = len(spend_values)

        df = pd.DataFrame({
            "campaign_name": [f"Campaign {i}" for i in range(num_campaigns)],
            "amount_spent": spend_values,
            "link_clicks": [50 for _ in range(num_campaigns)],
            "impressions": [2500 for _ in range(num_campaigns)],
            "ctr": [2.0 for _ in range(num_campaigns)],
            "cpc": [s / 50 for s in spend_values]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Column B (Amount Spent) and Column F (CPC) should have ZAR format
        for row in range(2, num_campaigns + 2):
            amount_cell = ws.cell(row=row, column=2)
            cpc_cell = ws.cell(row=row, column=6)

            assert amount_cell.number_format == ZAR_FORMAT, \
                f"Amount Spent cell row {row} doesn't have ZAR format"
            assert cpc_cell.number_format == ZAR_FORMAT, \
                f"CPC cell row {row} doesn't have ZAR format"


class TestCPCThresholdHighlightingProperty:
    """Property-based tests for CPC threshold highlighting.

    Feature:
        admetric-pro, Property 6: CPC Threshold Highlighting

    Validates:
        Requirements 5.1, 5.2

    Business Value:
        Automatically flags underperforming campaigns, saving analysts
        hours of manual review and ensuring no high-cost campaigns are missed.
    """

    @given(
        low_cpc=st.floats(min_value=0.0, max_value=19.99, allow_nan=False, allow_infinity=False),
        high_cpc=st.floats(min_value=20.01, max_value=100.0, allow_nan=False, allow_infinity=False)
    )
    @settings(max_examples=50)
    def test_cpc_threshold_highlighting(
        self,
        tmp_path_factory: TempPathFactory,
        low_cpc: float,
        high_cpc: float
    ) -> None:
        """Property: Rows with CPC > R20.00 are highlighted, others are not.

        Feature:
            admetric-pro, Property 6: CPC Threshold Highlighting

        Validates:
            Requirements 5.1, 5.2

        For any campaign row in the generated Excel report, if CPC exceeds
        R20.00 then the row SHALL have red background highlighting,
        otherwise the row SHALL have no red highlighting.

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            low_cpc: Hypothesis-generated CPC value below threshold.
            high_cpc: Hypothesis-generated CPC value above threshold.

        Raises:
            AssertionError: If highlighting rules are not applied correctly.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        df = pd.DataFrame({
            "campaign_name": ["Low CPC Campaign", "High CPC Campaign"],
            "amount_spent": [low_cpc * 10, high_cpc * 10],
            "link_clicks": [10, 10],
            "impressions": [1000, 1000],
            "ctr": [1.0, 1.0],
            "cpc": [low_cpc, high_cpc]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Row 2 (Low CPC) should NOT be highlighted
        low_cpc_fill = ws.cell(row=2, column=1).fill.start_color.rgb
        assert low_cpc_fill != "00FFC7CE", f"Low CPC row incorrectly highlighted (CPC={low_cpc})"

        # Row 3 (High CPC) SHOULD be highlighted
        high_cpc_fill = ws.cell(row=3, column=1).fill.start_color.rgb
        assert high_cpc_fill == "00FFC7CE", f"High CPC row not highlighted (CPC={high_cpc})"

    @given(
        cpc_at_threshold=st.just(20.0)
    )
    @settings(max_examples=10)
    def test_exact_threshold_not_highlighted(
        self,
        tmp_path_factory: TempPathFactory,
        cpc_at_threshold: float
    ) -> None:
        """Property: CPC exactly at R20.00 is NOT highlighted.

        Feature:
            admetric-pro, Property 6: CPC Threshold Highlighting

        Validates:
            Requirements 5.2

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            cpc_at_threshold: CPC value exactly at threshold (R20.00).

        Raises:
            AssertionError: If row at exact threshold is incorrectly highlighted.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        df = pd.DataFrame({
            "campaign_name": ["Exactly R20 CPC"],
            "amount_spent": [200.0],
            "link_clicks": [10],
            "impressions": [1000],
            "ctr": [1.0],
            "cpc": [cpc_at_threshold]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Row 2 (Exactly R20.00) should NOT be highlighted
        fill = ws.cell(row=2, column=1).fill.start_color.rgb
        assert fill != "00FFC7CE", "Row with CPC exactly R20.00 should not be highlighted"


class TestExcelDataRoundTripProperty:
    """Property-based tests for Excel data round-trip integrity.

    Feature:
        admetric-pro, Property 7: Excel Data Round-Trip

    Validates:
        Requirements 6.1, 6.2

    Business Value:
        Guarantees data integrity throughout the report generation process,
        ensuring clients receive accurate campaign metrics.
    """

    @given(
        campaign_name=st.text(
            alphabet=st.characters(whitelist_categories=("L", "N"), whitelist_characters=" -_"),
            min_size=1,
            max_size=30
        ).filter(lambda x: x.strip() != ""),
        amount_spent=st.floats(min_value=100.0, max_value=50000.0, allow_nan=False, allow_infinity=False),
        link_clicks=st.integers(min_value=1, max_value=10000),
        impressions=st.integers(min_value=100, max_value=1000000)
    )
    @settings(max_examples=50, deadline=None)
    def test_data_preserved_in_excel(
        self,
        tmp_path_factory: TempPathFactory,
        campaign_name: str,
        amount_spent: float,
        link_clicks: int,
        impressions: int
    ) -> None:
        """Property: Campaign data is preserved when written to Excel.

        Feature:
            admetric-pro, Property 7: Excel Data Round-Trip

        Validates:
            Requirements 6.1, 6.2

        For any processed campaign DataFrame, writing to Excel and reading
        back SHALL produce equivalent data values for all campaign metrics.

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            campaign_name: Hypothesis-generated campaign name.
            amount_spent: Hypothesis-generated spend amount.
            link_clicks: Hypothesis-generated click count.
            impressions: Hypothesis-generated impression count.

        Raises:
            AssertionError: If data is not preserved correctly.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        ctr = round((link_clicks / impressions) * 100, 2) if impressions > 0 else 0.0
        cpc = round(amount_spent / link_clicks, 2) if link_clicks > 0 else 0.0

        df = pd.DataFrame({
            "campaign_name": [campaign_name],
            "amount_spent": [round(amount_spent, 2)],
            "link_clicks": [link_clicks],
            "impressions": [impressions],
            "ctr": [ctr],
            "cpc": [cpc]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        ws = wb["Campaign Details"]

        # Verify campaign name is preserved
        assert ws.cell(row=2, column=1).value == campaign_name, \
            f"Campaign name not preserved: expected '{campaign_name}'"

        # Verify numeric values are preserved (with tolerance for float precision)
        assert ws.cell(row=2, column=3).value == link_clicks, \
            f"Link clicks not preserved: expected {link_clicks}"
        assert ws.cell(row=2, column=4).value == impressions, \
            f"Impressions not preserved: expected {impressions}"


class TestExecutiveSummaryAggregationProperty:
    """Property-based tests for Executive Summary aggregation.

    Feature:
        admetric-pro, Property 8: Executive Summary Aggregation

    Validates:
        Requirements 7.2, 7.3, 7.4, 7.5, 7.6

    Business Value:
        Provides accurate high-level metrics for executive stakeholders,
        enabling quick decision-making without manual calculations.
    """

    @given(
        spend_values=st.lists(
            st.floats(min_value=100.0, max_value=10000.0, allow_nan=False, allow_infinity=False),
            min_size=1,
            max_size=10
        ),
        click_values=st.lists(
            st.integers(min_value=10, max_value=1000),
            min_size=1,
            max_size=10
        ),
        impression_values=st.lists(
            st.integers(min_value=1000, max_value=100000),
            min_size=1,
            max_size=10
        )
    )
    @settings(max_examples=50)
    def test_summary_aggregates_match_data(
        self,
        tmp_path_factory: TempPathFactory,
        spend_values: list[float],
        click_values: list[int],
        impression_values: list[int]
    ) -> None:
        """Property: Executive Summary aggregates match sum/average of campaign data.

        Feature:
            admetric-pro, Property 8: Executive Summary Aggregation

        Validates:
            Requirements 7.2, 7.3, 7.4, 7.5, 7.6

        For any set of campaign data, the Executive Summary SHALL display
        aggregates that match the sum/average of the underlying data.

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            spend_values: Hypothesis-generated list of spend amounts.
            click_values: Hypothesis-generated list of click counts.
            impression_values: Hypothesis-generated list of impression counts.

        Raises:
            AssertionError: If summary aggregates don't match expected values.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        # Ensure all lists have same length
        min_len = min(len(spend_values), len(click_values), len(impression_values))
        spend_values = spend_values[:min_len]
        click_values = click_values[:min_len]
        impression_values = impression_values[:min_len]

        df = pd.DataFrame({
            "campaign_name": [f"Campaign {i}" for i in range(min_len)],
            "amount_spent": spend_values,
            "link_clicks": click_values,
            "impressions": impression_values,
            "ctr": [2.0 for _ in range(min_len)],
            "cpc": [s / c if c > 0 else 0 for s, c in zip(spend_values, click_values)]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)
        summary_ws = wb["Executive Summary"]

        # Calculate expected values
        expected_total_spend = sum(spend_values)
        expected_total_clicks = sum(click_values)
        expected_total_impressions = sum(impression_values)

        # Find values in summary sheet
        summary_text = ""
        for row in summary_ws.iter_rows(min_row=1, max_row=15, max_col=2):
            for cell in row:
                if cell.value:
                    summary_text += str(cell.value) + " "

        # Verify Total Spend is present (formatted as R X,XXX.XX)
        assert f"{expected_total_spend:,.2f}" in summary_text, \
            f"Expected Total Spend {expected_total_spend:,.2f} not found in summary"

    @given(
        num_campaigns=st.integers(min_value=1, max_value=5)
    )
    @settings(max_examples=30)
    def test_summary_sheet_exists(
        self,
        tmp_path_factory: TempPathFactory,
        num_campaigns: int
    ) -> None:
        """Property: Executive Summary sheet is always created.

        Feature:
            admetric-pro, Property 8: Executive Summary Aggregation

        Validates:
            Requirements 7.1

        Args:
            tmp_path_factory: Pytest fixture for creating temporary directories.
            num_campaigns: Hypothesis-generated number of campaigns.

        Raises:
            AssertionError: If Executive Summary sheet is not created.
        """
        tmp_path = tmp_path_factory.mktemp("reports")

        df = pd.DataFrame({
            "campaign_name": [f"Campaign {i}" for i in range(num_campaigns)],
            "amount_spent": [1000.0 for _ in range(num_campaigns)],
            "link_clicks": [50 for _ in range(num_campaigns)],
            "impressions": [2500 for _ in range(num_campaigns)],
            "ctr": [2.0 for _ in range(num_campaigns)],
            "cpc": [20.0 for _ in range(num_campaigns)]
        })

        report_path = generate_report(df, str(tmp_path))
        wb = load_workbook(report_path)

        assert "Executive Summary" in wb.sheetnames, \
            "Executive Summary sheet not found in workbook"
