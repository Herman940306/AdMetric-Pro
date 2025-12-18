"""Excel Formatter module for AdMetric Pro.

This module generates professionally formatted Excel reports from
campaign data, including styled headers, ZAR currency formatting,
conditional highlighting, and executive summary dashboards.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

# Configure module logger
logger = logging.getLogger(__name__)

# Style constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark Blue
HEADER_FONT = Font(bold=True, color="FFFFFF")  # White Bold
HIGH_CPC_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red
SUMMARY_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # Dashboard Blue
SUMMARY_VALUE_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # Light Gray

# Border style for dashboard
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ZAR currency format
ZAR_FORMAT = 'R #,##0.00'
PERCENTAGE_FORMAT = '0.00"%"'

# CPC threshold for conditional formatting (R20.00)
CPC_THRESHOLD = 20.00


def generate_timestamped_filename() -> str:
    """Generate a unique filename with current timestamp.

    Creates a filename following the format AdMetric_Pro_Report_YYYY-MM-DD_HHMM.xlsx
    to prevent overwriting previous reports and maintain report history.

    Returns:
        Filename string with current timestamp.

    Example:
        >>> filename = generate_timestamped_filename()
        >>> # Returns: "AdMetric_Pro_Report_2025-12-18_1430.xlsx"
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"AdMetric_Pro_Report_{timestamp}.xlsx"
    logger.info("Generated timestamped filename: %s", filename)
    return filename


def apply_header_formatting(worksheet: Worksheet, header_row: int = 1) -> None:
    """Apply professional styling to header row.

    Formats the header row with dark blue background and white bold text,
    creating a polished, client-ready appearance.

    Args:
        worksheet: The openpyxl worksheet to format.
        header_row: Row number containing headers (default: 1).

    Example:
        >>> apply_header_formatting(ws)
        # Headers now have dark blue background with white bold text
    """
    logger.debug("Applying header formatting to row %d", header_row)
    for cell in worksheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_currency_formatting(worksheet: Worksheet, columns: list[str], df: pd.DataFrame) -> None:
    """Apply ZAR currency format to specified columns.

    Formats monetary columns with South African Rand formatting (R #,##0.00)
    for professional financial presentation.

    Args:
        worksheet: The openpyxl worksheet to format.
        columns: List of column names to format as currency.
        df: DataFrame used to identify column positions.

    Example:
        >>> apply_currency_formatting(ws, ["amount_spent", "cpc"], df)
        # Spend and CPC columns now show as R 1,234.56
    """
    column_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}

    for col_name in columns:
        if col_name in column_indices:
            col_idx = column_indices[col_name]
            logger.debug("Applying ZAR format to column %s (index %d)", col_name, col_idx)
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = ZAR_FORMAT


def apply_percentage_formatting(worksheet: Worksheet, columns: list[str], df: pd.DataFrame) -> None:
    """Apply percentage format to specified columns.

    Formats percentage columns with two decimal places for clear
    metric presentation.

    Args:
        worksheet: The openpyxl worksheet to format.
        columns: List of column names to format as percentages.
        df: DataFrame used to identify column positions.

    Example:
        >>> apply_percentage_formatting(ws, ["ctr"], df)
        # CTR column now shows as 2.50%
    """
    column_indices = {col: idx + 1 for idx, col in enumerate(df.columns)}

    for col_name in columns:
        if col_name in column_indices:
            col_idx = column_indices[col_name]
            logger.debug("Applying percentage format to column %s", col_name)
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).number_format = PERCENTAGE_FORMAT


def apply_conditional_formatting(worksheet: Worksheet, cpc_column: str, df: pd.DataFrame, threshold: float = CPC_THRESHOLD) -> None:
    """Apply red highlighting to rows where CPC exceeds threshold.

    Highlights underperforming campaigns (CPC > R20.00) with light red
    background, making it easy for clients to identify areas needing attention.

    Args:
        worksheet: The openpyxl worksheet to format.
        cpc_column: Name of the CPC column.
        df: DataFrame containing the data.
        threshold: CPC threshold for highlighting (default: R20.00).

    Example:
        >>> apply_conditional_formatting(ws, "cpc", df)
        # Rows with CPC > R20.00 now have light red background
    """
    if cpc_column not in df.columns:
        logger.warning("CPC column '%s' not found, skipping conditional formatting", cpc_column)
        return

    cpc_col_idx = list(df.columns).index(cpc_column) + 1
    highlighted_count = 0

    for row_idx, cpc_value in enumerate(df[cpc_column], start=2):
        if cpc_value > threshold:
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col).fill = HIGH_CPC_FILL
            highlighted_count += 1

    logger.info("Highlighted %d campaigns with CPC > R%.2f", highlighted_count, threshold)


def create_executive_summary(workbook: Workbook, df: pd.DataFrame) -> Worksheet:
    """Create Executive Summary worksheet with aggregated metrics.

    Builds a high-end dashboard-style summary sheet displaying:
    - Total Spend (ZAR)
    - Total Impressions
    - Total Clicks
    - Average CPC (ZAR)
    - Overall CTR (%)

    Args:
        workbook: The openpyxl Workbook to add the summary to.
        df: DataFrame containing campaign data with metrics.

    Returns:
        The created Executive Summary worksheet.

    Example:
        >>> summary_ws = create_executive_summary(wb, df)
        # Creates a polished summary dashboard
    """
    logger.info("Creating Executive Summary sheet...")

    # Create the summary worksheet
    summary_ws = workbook.create_sheet(title="Executive Summary")

    # Calculate aggregate metrics
    total_spend = df["amount_spent"].sum() if "amount_spent" in df.columns else 0
    total_impressions = df["impressions"].sum() if "impressions" in df.columns else 0
    total_clicks = df["link_clicks"].sum() if "link_clicks" in df.columns else 0

    # Calculate averages with zero-division protection
    avg_cpc = total_spend / total_clicks if total_clicks > 0 else 0
    overall_ctr = (total_clicks / total_impressions) * 100 if total_impressions > 0 else 0

    # Summary data
    summary_data = [
        ("Total Spend", f"R {total_spend:,.2f}"),
        ("Total Impressions", f"{total_impressions:,}"),
        ("Total Clicks", f"{total_clicks:,}"),
        ("Average CPC", f"R {avg_cpc:.2f}"),
        ("Overall CTR", f"{overall_ctr:.2f}%"),
    ]

    # Title row
    summary_ws.merge_cells("A1:B1")
    title_cell = summary_ws["A1"]
    title_cell.value = "AdMetric Pro - Executive Summary"
    title_cell.font = Font(bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.row_dimensions[1].height = 30

    # Subtitle with report date
    summary_ws.merge_cells("A2:B2")
    date_cell = summary_ws["A2"]
    date_cell.value = f"Report Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    date_cell.font = Font(italic=True, size=10, color="666666")
    date_cell.alignment = Alignment(horizontal="center")
    summary_ws.row_dimensions[2].height = 20

    # Empty row for spacing
    summary_ws.row_dimensions[3].height = 15

    # Header row for metrics table
    summary_ws["A4"] = "Metric"
    summary_ws["B4"] = "Value"
    for cell in [summary_ws["A4"], summary_ws["B4"]]:
        cell.fill = SUMMARY_HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    summary_ws.row_dimensions[4].height = 25

    # Populate metrics
    for idx, (metric, value) in enumerate(summary_data, start=5):
        metric_cell = summary_ws.cell(row=idx, column=1, value=metric)
        value_cell = summary_ws.cell(row=idx, column=2, value=value)

        # Style metric name
        metric_cell.font = Font(bold=True)
        metric_cell.alignment = Alignment(horizontal="left", vertical="center")
        metric_cell.border = THIN_BORDER

        # Style value
        value_cell.fill = SUMMARY_VALUE_FILL
        value_cell.alignment = Alignment(horizontal="right", vertical="center")
        value_cell.border = THIN_BORDER

        # Highlight Total Spend and Avg CPC with currency styling
        if "Spend" in metric or "CPC" in metric:
            value_cell.font = Font(bold=True, color="1F4E79")

        summary_ws.row_dimensions[idx].height = 22

    # Set column widths
    summary_ws.column_dimensions["A"].width = 25
    summary_ws.column_dimensions["B"].width = 20

    logger.info("Executive Summary created: Total Spend R%.2f, %d campaigns", total_spend, len(df))

    return summary_ws


def auto_adjust_column_widths(worksheet: Worksheet) -> None:
    """Auto-adjust column widths based on content.

    Ensures all data is visible without manual column resizing,
    improving the professional appearance of the report.

    Args:
        worksheet: The openpyxl worksheet to adjust.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width


def generate_report(df: pd.DataFrame, output_dir: str = "output") -> str:
    """Generate a professionally formatted Excel report.

    Creates a complete Excel workbook with:
    - Campaign Details sheet with styled headers and conditional formatting
    - Executive Summary sheet with aggregated metrics dashboard

    This is the main entry point for report generation, orchestrating
    all formatting functions to produce client-ready output.

    Args:
        df: DataFrame with campaign data and calculated metrics.
            Expected columns: campaign_name, amount_spent, link_clicks,
            impressions, ctr, cpc.
        output_dir: Directory for output file (default: "output").

    Returns:
        Full path to the generated Excel file.

    Raises:
        IOError: If the file cannot be written.

    Example:
        >>> from src.csv_reader import read_meta_csv
        >>> from src.metrics import add_metrics_to_dataframe
        >>> df = read_meta_csv("data.csv")
        >>> df = add_metrics_to_dataframe(df)
        >>> report_path = generate_report(df)
        >>> print(f"Report saved to: {report_path}")
    """
    logger.info("Generating Excel report for %d campaigns...", len(df))

    # Create output directory if it doesn't exist
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Generate filename and full path
    filename = generate_timestamped_filename()
    file_path = output_path / filename

    try:
        # Create workbook
        workbook = Workbook()
        details_ws = workbook.active
        details_ws.title = "Campaign Details"

        # Prepare display columns with friendly names
        display_df = df.copy()
        display_df.columns = [
            "Campaign Name", "Amount Spent (ZAR)", "Link Clicks",
            "Impressions", "CTR (%)", "CPC (ZAR)"
        ]

        # Write data to Campaign Details sheet
        logger.info("Writing campaign data to worksheet...")
        for row_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                details_ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply formatting to Campaign Details
        apply_header_formatting(details_ws)

        # Map display names back to internal names for formatting functions
        internal_df = df.copy()
        apply_currency_formatting(details_ws, ["amount_spent", "cpc"], internal_df)
        apply_percentage_formatting(details_ws, ["ctr"], internal_df)
        apply_conditional_formatting(details_ws, "cpc", internal_df)
        auto_adjust_column_widths(details_ws)

        # Create Executive Summary sheet
        create_executive_summary(workbook, df)

        # Save workbook
        logger.info("Saving report to: %s", file_path)
        workbook.save(file_path)

        logger.info("Report generated successfully: %s", file_path)
        return str(file_path)

    except Exception as e:
        logger.error("Failed to generate report: %s", e)
        raise IOError(f"Failed to write Excel report: {e}")
